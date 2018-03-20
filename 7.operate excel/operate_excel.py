#coding:utf-8

import openpyxl
#创建表单对象
wb = openpyxl.load_workbook('example.xlsx')

#获取表单的名字
# print(wb.sheetnames)
#
# for sheet in wb:
#     print(sheet.title)

#创建一个表单
mysheet = wb.create_sheet('mysheet')

#获取一个表单对象
# mysheetobj = wb.get_sheet_by_name('mysheet')
#mysheetobj1 = wb['Sheet1']
mysheetobj1 = wb.active

#获取单元格
# c = mysheetobj1['A1']
# print(c.row)
# print(c.column)
# print(c.value)
# print(c.coordinate)
# print(mysheetobj1.cell(row=1,column=1).value)

#获取行和列
#
#
# col_range = mysheetobj1['B:C']
# row_range = mysheetobj1['1:3']
#
# for col in col_range:
#     for cell in col:
#         print(cell.value)
#
# for row in row_range:
#     for cell in row:
#         print(cell.value)

cell_range = mysheetobj1['A1:B3']

for rowOfCellObj  in cell_range:
    for cellObj in rowOfCellObj:
        print(cellObj.value)
    print('-'*10)


